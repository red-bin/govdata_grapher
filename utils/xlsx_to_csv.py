#!/usr/bin/python3

import csv
from openpyxl import load_workbook
from multiprocessing import Pool

def load_wb(fp):
    wb = load_workbook(fp, read_only=True, data_only=True)
    worksheets = wb.worksheets

    ret = []
    for ws in worksheets:
        rows_data = []
        for row in ws.iter_rows():
            row_data = [r.value for r in row]
            rows_data.append(row_data)

        print(fp, rows_data[0])
        if rows_data:
            ret += rows_data
            
    return (fp, ret)

file_list = """/opt/data/towing/ts1a towed.xlsx
/opt/data/towing/ts1b towed.xlsx
/opt/data/towing/ts2a color.xlsx
/opt/data/towing/ts2b color.xlsx
/opt/data/towing/ts3a VIN.xlsx
/opt/data/towing/ts3b VIN.xlsx
/opt/data/towing/ts4a reason.xlsx
/opt/data/towing/ts4b reason.xlsx
/opt/data/towing/ts5a contract.xlsx
/opt/data/towing/ts6a make.xlsx
/opt/data/towing/ts5b contract.xlsx
/opt/data/towing/ts6b make.xlsx""".split('\n')

pool = Pool(processes=len(file_list))


#loaded = [(fp, load_wb(fp)) for fp in file_list)

pre_loaded = pool.map(load_wb, file_list)

for fp, data in pre_loaded:
    new_fp = fp.replace('.xlsx', '.csv')
    fh = open(new_fp, 'w') 

    w = csv.writer(fh)
    w.writerows(data)
    fh.close()
